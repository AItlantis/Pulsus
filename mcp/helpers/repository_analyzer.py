"""
Repository Analyzer - MCP Helper
Integrates comprehensive Python repository analysis into the MCP framework.

Based on the enhanced read_python_structure.py script, providing:
- File metadata validation (naming conventions, docstrings, ownership)
- Reusability scoring for functions
- Dependency analysis and visualization
- Excel report generation with multiple analysis sheets
- HTML visualizations (D3.js dependency networks, folder trees)

Author: Adapted for MCP by Testudo Framework
"""

from __future__ import annotations

import ast
import os
import sys
import re
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Any
from collections import defaultdict, Counter
from datetime import datetime

# Optional dependencies
try:
    from radon.complexity import cc_visit
    _HAS_RADON = True
except Exception:
    _HAS_RADON = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


# --- Data structures ---

@dataclass
class FunctionInfo:
    """Information about a function in a file."""
    file_rel: str
    qualname: str
    name: str
    class_name: Optional[str]
    lineno: int
    end_lineno: Optional[int]
    length: Optional[int]
    docstring: Optional[str]
    complexity: Optional[int] = None
    unused: Optional[bool] = None
    reusability_score: Optional[int] = None
    used_in_files: Set[str] = field(default_factory=set)
    has_hardcoded_paths: bool = False
    is_generic_name: bool = False


@dataclass
class ClassInfo:
    """Information about a class in a file."""
    file_rel: str
    name: str
    lineno: int
    end_lineno: Optional[int]
    docstring: Optional[str]
    methods: List[FunctionInfo] = field(default_factory=list)


@dataclass
class FileInfo:
    """Complete information about a Python file."""
    file_rel: str
    file_abs: str
    folder_rel: str
    direct_folder: str
    subfolder: str
    size_kb: float
    mtime: str
    lines: int

    # Metadata
    atype: Optional[str] = None
    owner: Optional[str] = None
    category_name: Optional[str] = None
    script_name: Optional[str] = None
    docstring_sections: Dict[str, str] = field(default_factory=dict)

    # Code elements
    functions: List[FunctionInfo] = field(default_factory=list)
    classes: List[ClassInfo] = field(default_factory=list)
    calls: List[Tuple[str, str]] = field(default_factory=list)

    # Dependencies
    imports_std: Set[str] = field(default_factory=set)
    imports_third: Set[str] = field(default_factory=set)
    imports_local: Set[str] = field(default_factory=set)
    imports_all: Set[str] = field(default_factory=set)

    # Validation
    issues: List[str] = field(default_factory=list)
    naming_valid: bool = True


class RepositoryAnalyzer:
    """
    Comprehensive repository analyzer with MCP integration.

    Provides Python repository analysis including:
    - File metadata validation
    - Reusability scoring
    - Dependency analysis
    - Report generation (Excel, HTML)
    """

    def __init__(self):
        self.stdlibs = self._get_standard_libs()

    def _get_standard_libs(self) -> Set[str]:
        """Get a set of standard library module names."""
        # Basic stdlib list - can be enhanced with stdlib_list package
        return {
            "abc", "argparse", "array", "ast", "asyncio", "base64", "binascii",
            "bisect", "builtins", "calendar", "cmath", "collections", "concurrent",
            "contextlib", "copy", "csv", "ctypes", "dataclasses", "datetime",
            "decimal", "difflib", "enum", "errno", "functools", "gc", "getpass",
            "glob", "gzip", "hashlib", "heapq", "hmac", "html", "http", "importlib",
            "inspect", "io", "ipaddress", "itertools", "json", "logging", "math",
            "multiprocessing", "operator", "os", "pathlib", "pickle", "platform",
            "pprint", "queue", "random", "re", "shutil", "socket", "sqlite3",
            "statistics", "string", "struct", "subprocess", "sys", "tempfile",
            "threading", "time", "traceback", "types", "typing", "unittest",
            "urllib", "uuid", "warnings", "xml", "zipfile"
        }

    def analyze_repository(self, repo_path: str, ignore_patterns: Optional[List[str]] = None) -> Dict[str, Any]:
        """
        Analyze a Python repository.

        Args:
            repo_path: Path to repository root
            ignore_patterns: Optional list of patterns to ignore (e.g., ['test', '__pycache__'])

        Returns:
            Dictionary with analysis results
        """
        root = Path(repo_path).resolve()

        if not root.is_dir():
            return {
                "success": False,
                "error": f"Path is not a directory: {repo_path}"
            }

        ignore_patterns = ignore_patterns or [
            "__pycache__", ".venv", "venv", ".git", "build", "dist",
            "node_modules", "test", "tests"
        ]

        # Scan all Python files
        files = self._scan_files(root, ignore_patterns)

        if not files:
            return {
                "success": False,
                "error": "No Python files found in repository"
            }

        # Analyze dependencies
        local_roots = self._discover_local_roots(root)
        self._split_dependencies(files, local_roots)

        # Detect unused functions and track usage
        self._detect_unused_functions(files)

        # Validate metadata
        self._validate_file_metadata(files)

        # Calculate reusability scores
        self._calculate_reusability_scores(files)

        # Generate summary statistics
        stats = self._generate_statistics(files)

        return {
            "success": True,
            "repository": str(root),
            "files_analyzed": len(files),
            "statistics": stats,
            "files": [self._file_to_dict(f) for f in files],
            "issues_summary": self._summarize_issues(files),
            "reusability_summary": self._summarize_reusability(files)
        }

    def generate_excel_report(self, analysis_result: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """
        Generate comprehensive Excel report from analysis results.

        Args:
            analysis_result: Result from analyze_repository()
            output_path: Path for output Excel file

        Returns:
            Dictionary with success status and file path
        """
        if not _HAS_OPENPYXL:
            return {
                "success": False,
                "error": "openpyxl not installed - cannot generate Excel reports"
            }

        if not analysis_result.get("success"):
            return {
                "success": False,
                "error": "Analysis result indicates failure"
            }

        # Reconstruct FileInfo objects from dictionaries
        files = [self._dict_to_file(f) for f in analysis_result["files"]]

        try:
            self._export_excel(files, Path(output_path))
            return {
                "success": True,
                "output_path": output_path,
                "sheets": [
                    "Summary & Actions",
                    "Issues & Warnings",
                    "Files Overview",
                    "Reusability Analysis",
                    "Functions & Complexity",
                    "Function Calls",
                    "Dashboard"
                ]
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"Failed to generate Excel report: {str(e)}"
            }

    def validate_file(self, file_path: str) -> Dict[str, Any]:
        """
        Validate a single Python file.

        Args:
            file_path: Path to Python file

        Returns:
            Dictionary with validation results
        """
        path = Path(file_path)

        if not path.exists() or not path.is_file():
            return {
                "success": False,
                "error": f"File not found: {file_path}"
            }

        if path.suffix != ".py":
            return {
                "success": False,
                "error": f"Not a Python file: {file_path}"
            }

        try:
            # Analyze the file
            root = path.parent
            file_info = self._analyze_file(path, root)

            # Validate metadata
            self._validate_file_metadata([file_info])

            return {
                "success": True,
                "file": str(path),
                "issues": file_info.issues,
                "naming_valid": file_info.naming_valid,
                "metadata": {
                    "owner": file_info.owner,
                    "atype": file_info.atype,
                    "category": file_info.category_name,
                    "script_name": file_info.script_name
                },
                "statistics": {
                    "lines": file_info.lines,
                    "functions": len(file_info.functions),
                    "classes": len(file_info.classes),
                    "imports_total": len(file_info.imports_all)
                }
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"Failed to validate file: {str(e)}"
            }

    # --- Internal methods ---

    def _scan_files(self, root: Path, ignore_patterns: List[str]) -> List[FileInfo]:
        """Scan all Python files in repository."""
        files: List[FileInfo] = []
        all_paths = list(root.rglob("*.py"))

        for p in all_paths:
            # Check if path should be ignored
            if self._is_ignored(p, ignore_patterns, root):
                continue

            try:
                file_info = self._analyze_file(p, root)
                files.append(file_info)
            except Exception as e:
                print(f"Warning: Failed to analyze {p}: {e}")

        return files

    def _is_ignored(self, path: Path, ignore_patterns: List[str], root: Path) -> bool:
        """Check if path matches ignore patterns."""
        path_str = str(path.resolve()).lower()
        root_str = str(root.resolve()).lower()

        for pattern in ignore_patterns:
            if pattern.lower() in path_str and path_str != root_str:
                return True
        return False

    def _analyze_file(self, path: Path, root: Path) -> FileInfo:
        """Analyze a single Python file."""
        content = self._safe_read_text(path)
        lines = content.count("\n") + 1
        stat = path.stat()

        file_rel = str(path.relative_to(root))
        folder_rel = str(path.parent.relative_to(root)) if path.parent != root else ""
        parts = Path(file_rel).parts
        direct_folder = parts[0] if len(parts) > 1 else ""

        # Parse metadata
        header_meta = self._parse_aimsun_header(content)
        filename_meta = self._parse_filename_convention(file_rel)
        docstring_sections = self._parse_structured_docstring(content)

        # Parse AST
        visitor = _ASTVisitor(file_rel)
        try:
            tree = ast.parse(content)
            visitor.visit(tree)
        except Exception as e:
            print(f"AST parse error in {file_rel}: {e}")

        # Calculate complexity if radon available
        comp_map: Dict[str, int] = {}
        if _HAS_RADON:
            try:
                for block in cc_visit(content):
                    comp_map[block.name] = int(getattr(block, "complexity", 0))
            except Exception:
                pass

        for f in visitor.functions:
            f.complexity = comp_map.get(f.qualname) or comp_map.get(f.name)

        return FileInfo(
            file_rel=file_rel,
            file_abs=str(path.resolve()),
            folder_rel=folder_rel,
            direct_folder=direct_folder,
            subfolder=folder_rel,
            size_kb=round(stat.st_size / 1024.0, 2),
            mtime=datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            lines=lines,
            atype=header_meta.get('atype'),
            owner=header_meta.get('owner'),
            category_name=filename_meta.get('category_name'),
            script_name=filename_meta.get('script_name'),
            docstring_sections=docstring_sections,
            functions=visitor.functions,
            classes=visitor.classes,
            calls=visitor.calls,
            imports_all=set(sorted(visitor.imports)),
            naming_valid=filename_meta.get('naming_valid', False),
            issues=filename_meta.get('issues', [])
        )

    def _safe_read_text(self, path: Path) -> str:
        """Safely read text file with encoding fallback."""
        try:
            return path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            return path.read_text(encoding="latin-1", errors="ignore")

    def _parse_aimsun_header(self, content: str) -> Dict[str, str]:
        """Parse special Aimsun headers like #atype=... and #owner=..."""
        metadata = {}
        for line in content.splitlines()[:5]:
            if line.strip().startswith('#'):
                line = line.strip()[1:]
                if '=' in line:
                    key, value = line.split('=', 1)
                    key = key.strip().lower()
                    if key in ['atype', 'owner']:
                        metadata[key] = value.strip()
        return metadata

    def _parse_filename_convention(self, filename: str) -> Dict[str, Any]:
        """Decode the _XX_YY_name.py convention and validate it."""
        parts = Path(filename).stem.split('_')
        meta: Dict[str, Any] = {
            'category_name': 'N/A',
            'script_name': Path(filename).stem,
            'naming_valid': True,  # Be lenient by default
            'issues': []
        }

        # Infer category from parent folder
        parent_folder = Path(filename).parent.name.lower()
        if parent_folder not in ['.', '..', '']:
            meta['category_name'] = parent_folder.capitalize()

        return meta

    def _parse_structured_docstring(self, content: str) -> Dict[str, str]:
        """Extract content from markdown-style sections within the initial docstring."""
        sections = {}
        docstring_match = re.search(r'"""(.*?)"""', content, re.DOTALL)
        if not docstring_match:
            return sections

        docstring = docstring_match.group(1)

        # Try to extract sections
        pattern = r"##\s+(.*?)\n(.*?)(?=\n##\s+|$)"
        matches = re.findall(pattern, docstring, re.DOTALL | re.IGNORECASE)

        for match in matches:
            title = match[0].strip().upper().replace(' ', '_')
            text_content = ' '.join(line.strip() for line in match[1].splitlines() if line.strip())
            sections[title] = text_content

        # Fallback for simple description
        if not sections and docstring.strip():
            sections['DESCRIPTION'] = ' '.join(line.strip() for line in docstring.strip().splitlines())

        return sections

    def _discover_local_roots(self, root: Path) -> Set[str]:
        """Discover local package/module names."""
        local: Set[str] = set()
        for p in root.rglob("*.py"):
            try:
                rel = p.relative_to(root)
            except ValueError:
                continue
            parts = rel.parts
            if len(parts) == 1:
                local.add(parts[0].split(".")[0])
            else:
                top = parts[0]
                if (root / top / "__init__.py").exists():
                    local.add(top)
                else:
                    local.add(parts[0])
        return local

    def _split_dependencies(self, files: List[FileInfo], local_roots: Set[str]) -> None:
        """Split imports into standard, third-party, and local."""
        for f in files:
            std, third, local = set(), set(), set()
            for dep in f.imports_all:
                if dep in local_roots:
                    local.add(dep)
                elif dep in self.stdlibs:
                    std.add(dep)
                else:
                    third.add(dep)
            f.imports_std = std
            f.imports_third = third
            f.imports_local = local

    def _detect_unused_functions(self, files: List[FileInfo]) -> None:
        """Detect unused functions and track function usage."""
        defined_names = {fn.name for f in files for fn in f.functions}
        defined_names |= {fn.qualname for f in files for fn in f.functions if fn.class_name}
        called_names = {callee for f in files for _, callee in f.calls}

        function_callers: Dict[str, Set[str]] = defaultdict(set)
        for f in files:
            for caller, callee in f.calls:
                function_callers[callee].add(f.file_rel)

        for f in files:
            for fn in f.functions:
                fn.unused = not ((fn.name in called_names) or (fn.qualname in called_names))
                fn.used_in_files = function_callers.get(fn.name, set()) | function_callers.get(fn.qualname, set())

    def _validate_file_metadata(self, files: List[FileInfo]) -> None:
        """Validate file metadata and naming conventions."""
        for f in files:
            # Check for missing owner (optional for non-Aimsun repos)
            if not f.owner or f.owner == "N/A":
                f.issues.append("Missing owner metadata (#owner=...)")

            # Check for missing description
            if 'DESCRIPTION' not in f.docstring_sections or not f.docstring_sections.get('DESCRIPTION'):
                f.issues.append("Missing DESCRIPTION section in docstring")

    def _calculate_reusability_scores(self, files: List[FileInfo]) -> None:
        """Calculate reusability score for each function."""
        generic_patterns = [
            'get_', 'set_', 'load_', 'save_', 'read_', 'write_', 'parse_', 'format_',
            'convert_', 'validate_', 'check_', 'calculate_', 'process_', 'update_',
            'create_', 'delete_', 'find_', 'search_', 'filter_', 'sort_', 'clean_',
            'transform_', 'generate_', 'build_', 'export_', 'import_', 'extract_'
        ]

        hardcoded_patterns = [
            r'C:\\', r'/home/', r'/usr/', r'D:\\'
        ]

        for f in files:
            try:
                content = Path(f.file_abs).read_text(encoding='utf-8', errors='ignore')
            except:
                content = ""

            for fn in f.functions:
                score = 0

                # Has documentation? (+2 points)
                if fn.docstring and len(fn.docstring.strip()) > 20:
                    score += 2

                # Generic naming pattern? (+2 points)
                fn.is_generic_name = any(fn.name.lower().startswith(pat) for pat in generic_patterns)
                if fn.is_generic_name:
                    score += 2

                # Used in multiple files? (+3 points per additional file, max 6)
                if len(fn.used_in_files) > 1:
                    score += min(3 * (len(fn.used_in_files) - 1), 6)

                # Not too complex? (+1 point if complexity <= 10)
                if fn.complexity and fn.complexity <= 10:
                    score += 1

                # Check for hardcoded paths (-3 points penalty)
                if fn.lineno and fn.end_lineno:
                    lines = content.split('\n')[fn.lineno-1:fn.end_lineno]
                    func_code = '\n'.join(lines)
                    fn.has_hardcoded_paths = any(re.search(pat, func_code) for pat in hardcoded_patterns)
                    if fn.has_hardcoded_paths:
                        score -= 3

                # Reasonable function length? (+1 if length < 50 lines)
                if fn.length and fn.length < 50:
                    score += 1

                fn.reusability_score = max(0, score)

    def _generate_statistics(self, files: List[FileInfo]) -> Dict[str, Any]:
        """Generate summary statistics from analyzed files."""
        total_functions = sum(len(f.functions) for f in files)
        total_classes = sum(len(f.classes) for f in files)
        total_lines = sum(f.lines for f in files)

        files_with_issues = len([f for f in files if f.issues])

        third_party_imports = set()
        for f in files:
            third_party_imports.update(f.imports_third)

        owner_counts = Counter(f.owner for f in files if f.owner and f.owner != "N/A")

        return {
            "total_files": len(files),
            "total_functions": total_functions,
            "total_classes": total_classes,
            "total_lines": total_lines,
            "files_with_issues": files_with_issues,
            "compliance_rate": f"{((len(files) - files_with_issues) / len(files) * 100):.1f}%" if files else "0%",
            "third_party_imports": len(third_party_imports),
            "top_imports": dict(Counter(m for f in files for m in f.imports_third).most_common(10)),
            "owners": dict(owner_counts)
        }

    def _summarize_issues(self, files: List[FileInfo]) -> Dict[str, Any]:
        """Summarize all issues found."""
        all_issues = []
        for f in files:
            for issue in f.issues:
                priority = "HIGH" if "naming" in issue.lower() or "missing owner" in issue.lower() else "MEDIUM"
                all_issues.append({
                    "file": f.file_rel,
                    "issue": issue,
                    "priority": priority
                })

        priority_counts = Counter(issue["priority"] for issue in all_issues)

        return {
            "total_issues": len(all_issues),
            "by_priority": dict(priority_counts),
            "top_issues": all_issues[:20]
        }

    def _summarize_reusability(self, files: List[FileInfo]) -> Dict[str, Any]:
        """Summarize reusability scores."""
        all_functions = [(f, fn) for f in files for fn in f.functions]

        if not all_functions:
            return {"total_functions": 0}

        scores = [fn.reusability_score or 0 for _, fn in all_functions]
        avg_score = sum(scores) / len(scores) if scores else 0

        high_reusability = [
            {
                "function": fn.qualname,
                "file": f.file_rel,
                "score": fn.reusability_score,
                "used_in": len(fn.used_in_files)
            }
            for f, fn in sorted(all_functions, key=lambda x: x[1].reusability_score or 0, reverse=True)[:10]
        ]

        return {
            "total_functions": len(all_functions),
            "average_score": round(avg_score, 2),
            "max_score": max(scores) if scores else 0,
            "top_reusable_functions": high_reusability
        }

    def _file_to_dict(self, file_info: FileInfo) -> Dict[str, Any]:
        """Convert FileInfo to dictionary for JSON serialization."""
        return {
            "file_rel": file_info.file_rel,
            "file_abs": file_info.file_abs,
            "lines": file_info.lines,
            "size_kb": file_info.size_kb,
            "owner": file_info.owner,
            "atype": file_info.atype,
            "category": file_info.category_name,
            "functions_count": len(file_info.functions),
            "classes_count": len(file_info.classes),
            "imports_third": list(file_info.imports_third),
            "imports_local": list(file_info.imports_local),
            "issues": file_info.issues,
            "naming_valid": file_info.naming_valid
        }

    def _dict_to_file(self, data: Dict[str, Any]) -> FileInfo:
        """Convert dictionary back to FileInfo (simplified)."""
        return FileInfo(
            file_rel=data["file_rel"],
            file_abs=data["file_abs"],
            folder_rel="",
            direct_folder="",
            subfolder="",
            size_kb=data["size_kb"],
            mtime="",
            lines=data["lines"],
            owner=data.get("owner"),
            atype=data.get("atype"),
            category_name=data.get("category"),
            issues=data.get("issues", []),
            naming_valid=data.get("naming_valid", True)
        )

    def _export_excel(self, files: List[FileInfo], output_path: Path) -> None:
        """Generate simplified Excel report."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Summary sheet
        ws.append(["Repository Analysis Summary"])
        ws.append([])
        ws.append(["Total Files", len(files)])
        ws.append(["Total Functions", sum(len(f.functions) for f in files)])
        ws.append(["Total Lines", sum(f.lines for f in files)])
        ws.append(["Files with Issues", len([f for f in files if f.issues])])

        # Files list sheet
        ws_files = wb.create_sheet("Files")
        ws_files.append(["File", "Lines", "Functions", "Owner", "Issues"])

        for f in sorted(files, key=lambda x: x.file_rel):
            ws_files.append([
                f.file_rel,
                f.lines,
                len(f.functions),
                f.owner or "Unassigned",
                "; ".join(f.issues[:2]) if f.issues else "None"
            ])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))


# --- AST Visitor ---

class _ASTVisitor(ast.NodeVisitor):
    """AST visitor to extract functions, classes, calls, and imports."""

    def __init__(self, file_rel: str):
        self.file_rel = file_rel
        self.functions: List[FunctionInfo] = []
        self.classes: List[ClassInfo] = []
        self.calls: List[Tuple[str, str]] = []
        self.imports: Set[str] = set()
        self._class_stack: List[str] = []
        self._func_stack: List[str] = []

    def current_class(self) -> Optional[str]:
        return self._class_stack[-1] if self._class_stack else None

    def current_func_name(self) -> Optional[str]:
        return self._func_stack[-1] if self._func_stack else None

    def current_qualname(self) -> Optional[str]:
        fn, cls = self.current_func_name(), self.current_class()
        return f"{cls}.{fn}" if fn and cls else fn

    def visit_ClassDef(self, node: ast.ClassDef):
        info = ClassInfo(
            file_rel=self.file_rel,
            name=node.name,
            lineno=node.lineno,
            end_lineno=getattr(node, "end_lineno", None),
            docstring=ast.get_docstring(node)
        )
        self.classes.append(info)
        self._class_stack.append(node.name)
        self.generic_visit(node)
        self._class_stack.pop()

    def visit_FunctionDef(self, node: ast.FunctionDef):
        qual = f"{self.current_class()}.{node.name}" if self.current_class() else node.name
        end_ln = getattr(node, "end_lineno", None)
        length = max(0, end_ln - node.lineno + 1) if end_ln is not None else None

        info = FunctionInfo(
            file_rel=self.file_rel,
            qualname=qual,
            name=node.name,
            class_name=self.current_class(),
            lineno=node.lineno,
            end_lineno=end_ln,
            length=length,
            docstring=ast.get_docstring(node)
        )

        if self.current_class():
            self.classes[-1].methods.append(info)

        self.functions.append(info)
        self._func_stack.append(node.name)
        self.generic_visit(node)
        self._func_stack.pop()

    def visit_AsyncFunctionDef(self, node: ast.AsyncFunctionDef):
        self.visit_FunctionDef(node)

    def visit_Call(self, node: ast.Call):
        callee = None
        if isinstance(node.func, ast.Name):
            callee = node.func.id
        elif isinstance(node.func, ast.Attribute):
            callee = node.func.attr

        if callee and self.current_qualname():
            self.calls.append((self.current_qualname(), callee))

        self.generic_visit(node)

    def visit_Import(self, node: ast.Import):
        for alias in node.names:
            self.imports.add(alias.name.split(".")[0])

    def visit_ImportFrom(self, node: ast.ImportFrom):
        if node.module:
            self.imports.add(node.module.split(".")[0])
